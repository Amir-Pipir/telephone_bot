import openpyxl as op

file_name = '12.11.23 p-e.xlsx'


async def check_for_admin(user_id):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    max_row2 = sheet2.max_row + 1
    admins = [sheet2[f'A{x}'].value for x in range(2, max_row2) if sheet2[f'D{x}'].value == 'admin']
    if user_id in admins:
        return True
    else:
        return False


async def find_us(us_id):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    for i in range(2, sheet2.max_row + 1):
        val = sheet2.cell(row=i, column=1).value
        if val == us_id:
            return sheet2.cell(row=i, column=2).value, i
    return False


async def admin(row):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    sheet2[f'D{row}'] = 'admin'
    wb2.save('Таблица пользователей.xlsx')


async def lock_unlock(us_id, status):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    max_row2 = sheet2.max_row + 1
    for i in range(2, max_row2):
        val = sheet2.cell(row=i, column=1).value
        if str(val) == us_id:
            sheet2[f"E{i}"] = status
            wb2.save('Таблица пользователей.xlsx')
            return sheet2.cell(row=i, column=2).value
    wb2.save('Таблица пользователей.xlsx')
    return False


async def check_users(us_id):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    max_row2 = sheet2.max_row + 1
    for i in range(2, max_row2):
        val = sheet2.cell(row=i, column=1).value
        if val == us_id:
            if str(sheet2[f'E{i}'].value) == "Разблок":
                return True
            else:
                return False


async def new_user(us_id, username, phone_number, role, status):
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    new_data = [us_id, username, phone_number, role, status]
    sheet2.append(new_data)
    wb2.save('Таблица пользователей.xlsx')


async def excel_read(m):
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_row = sheet.max_row + 1

    first_filter = []
    second_filter = []
    third_filter = []

    for i in range(2, max_row):
        val = sheet.cell(row=i, column=1).value
        if val:
            if m[0] in val.lower():
                first_filter.append([cell.value for cell in sheet[i]][0])
                first_filter.append([cell.value for cell in sheet[i]][1])

    if len(m) > 1:
        for x in range(0, len(first_filter), 2):
            if m[1] in first_filter[x].lower():
                second_filter.append(first_filter[x])
                second_filter.append(first_filter[x + 1])
        if len(m) > 2:
            for d in range(0, len(second_filter), 2):
                if m[2] in second_filter[d].lower():
                    third_filter.append(second_filter[d + 1])
            return third_filter
        return second_filter[1::2]
    return first_filter[1::2]


async def find_info(val):
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    result = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if row[2 - 1].value:
            if val[:62] == row[2 - 1].value[:62]:
                result.append([cell.value for cell in row[1:]])
    return result[0]


async def users_rows():
    wb2 = op.load_workbook('Таблица пользователей.xlsx')
    sheet2 = wb2.active
    max_row2 = sheet2.max_row
    return max_row2 - 1
